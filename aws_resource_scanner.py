import boto3
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def get_all_paginated_items(client, method_name, list_key, **kwargs):
    """
    Função auxiliar para lidar com a paginação de APIs do Boto3.
    Retorna uma lista consolidada de todos os itens.
    """
    paginator = client.get_paginator(method_name)
    all_items = []
    try:
        pages = paginator.paginate(**kwargs)
        for page in pages:
            all_items.extend(page.get(list_key, []))
    except Exception as e:
        print(f"    Erro de paginação para {method_name}: {e}")
    return all_items

def export_to_excel(data, filename="resources.xlsx"):
    """
    Exporta os dados coletados para um arquivo Excel.
    Cada tipo de recurso será uma aba diferente na planilha.
    """
    if not data:
        print("Nenhum dado para exportar para Excel.")
        return

    workbook = openpyxl.Workbook()
    
    # Remove a aba padrão criada automaticamente
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    print(f"\nExportando dados para {filename}...")

    for resource_type, resources_list in data.items():
        # Ignora chaves que não são listas ou estão vazias
        if not isinstance(resources_list, list) or not resources_list:
            continue

        # Cria uma nova aba para cada tipo de recurso
        # Limita o nome da aba a 31 caracteres, que é o limite do Excel
        sheet_name = str(resource_type)[:31]
        if sheet_name in workbook.sheetnames:
            # Garante nomes de abas únicos se houver colisão
            i = 1
            while f"{sheet_name[:28]}{i}" in workbook.sheetnames:
                i += 1
            sheet_name = f"{sheet_name[:28]}{i}"
        sheet = workbook.create_sheet(title=sheet_name)

        # Coleta todas as chaves (cabeçalhos) de todos os dicionários na lista de recursos
        all_keys = set()
        for res in resources_list:
            if isinstance(res, dict):
                all_keys.update(res.keys())
        
        headers = sorted(list(all_keys)) # Ordena os cabeçalhos para consistência
        if not headers: # Se não houver cabeçalhos, pula
            print(f"  Pulando '{resource_type}' (sem cabeçalhos para exportar).")
            workbook.remove(sheet) # Remove a aba vazia
            continue
            
        # Escreve os cabeçalhos na primeira linha
        sheet.append(headers)

        # Escreve os dados
        for resource in resources_list:
            row_data = []
            for header in headers:
                value = resource.get(header, '') # Pega o valor, ou string vazia se não existir
                
                # Trata objetos datetime para strings
                if isinstance(value, datetime):
                    row_data.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                # Trata listas ou dicionários aninhados convertendo para string JSON
                elif isinstance(value, (list, dict)):
                    try:
                        row_data.append(json.dumps(value, default=str, ensure_ascii=False))
                    except TypeError: # Fallback caso json.dumps falhe por algum motivo
                        row_data.append(str(value))
                else:
                    row_data.append(value)
            sheet.append(row_data)
        
        # Ajusta a largura das colunas automaticamente (opcional, pode ser lento para muitas colunas/linhas)
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in sheet[column_letter]:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        print(f"  Aba '{resource_type}' exportada com {len(resources_list)} itens.")

    try:
        workbook.save(filename)
        print(f"Relatório Excel salvo como '{filename}'.")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}: Verifique permissões ou se o arquivo não está aberto.")

def get_aws_resources_info():
    """
    Coleta e retorna informações sobre diversos recursos AWS na conta,
    incluindo paginação para serviços com muitos recursos.
    """
    resource_info = {}

    # --- AWS Organizations and OUs ---
    organizations_client = boto3.client('organizations')
    print("Coletando informações do AWS Organizations (Root, OUs e Contas)...")
    try:
        # Obter o Root da Organização
        roots = get_all_paginated_items(organizations_client, 'list_roots', 'Roots')
        all_org_info = []
        if roots:
            root = roots[0] # Uma organização tem apenas um Root
            root_id = root['Id']
            all_org_info.append({
                'Type': 'Root',
                'Id': root['Id'],
                'Arn': root['Arn'],
                'Name': root['Name'],
                'PolicyTypes': [{'Type': pt['Type'], 'Status': pt['Status']} for pt in root.get('PolicyTypes', [])]
            })
            print(f"  Encontrado Root da Organização: {root['Name']} ({root['Id']}).")

            # Função auxiliar recursiva para listar OUs e contas
            def list_ous_and_accounts(parent_id, parent_name="Root", level=0):
                # Listar OUs filhas
                ous = get_all_paginated_items(organizations_client, 'list_organizational_units_for_parent', 'OrganizationalUnits', ParentId=parent_id)
                for ou in ous:
                    all_org_info.append({
                        'Type': 'OrganizationalUnit',
                        'Id': ou['Id'],
                        'Arn': ou['Arn'],
                        'Name': ou['Name'],
                        'ParentId': parent_id,
                        'ParentName': parent_name,
                        'Level': level
                    })
                    print(f"    {'  ' * level}OU: {ou['Name']} ({ou['Id']})")
                    # Chamada recursiva para OUs aninhadas
                    list_ous_and_accounts(ou['Id'], ou['Name'], level + 1)
                
                # Listar contas filhas (diretamente sob o Root ou uma OU)
                accounts = get_all_paginated_items(organizations_client, 'list_accounts_for_parent', 'Accounts', ParentId=parent_id)
                for account in accounts:
                    all_org_info.append({
                        'Type': 'Account',
                        'Id': account['Id'],
                        'Arn': account['Arn'],
                        'Name': account['Name'],
                        'Email': account.get('Email'),
                        'Status': account['Status'],
                        'JoinedMethod': account['JoinedMethod'],
                        'JoinedTimestamp': account['JoinedTimestamp'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(account['JoinedTimestamp'], datetime) else str(account['JoinedTimestamp']),
                        'ParentId': parent_id,
                        'ParentName': parent_name,
                        'Level': level
                    })
                    print(f"    {'  ' * level}  Conta: {account['Name']} ({account['Id']})")

            list_ous_and_accounts(root_id)
        else:
            print("  Nenhuma organização AWS encontrada ou esta conta não é a conta de gerenciamento.")
        
        resource_info['AWSOrganizationStructure'] = all_org_info

    except Exception as e:
        print(f"  Erro ao listar recursos do AWS Organizations: {e}")
        resource_info['AWSOrganizationStructure'] = f"Erro ao listar: {e}"

    # --- EC2 Resources (EC2, VPCs, Transit Gateways) ---
    ec2_client = boto3.client('ec2')

    print("Coletando informações de instâncias EC2...")
    try:
        instances = get_all_paginated_items(ec2_client, 'describe_instances', 'Reservations')
        all_instances = []
        for reservation in instances:
            for instance in reservation['Instances']:
                all_instances.append({
                    'InstanceId': instance['InstanceId'],
                    'InstanceType': instance['InstanceType'],
                    'State': instance['State']['Name'],
                    'LaunchTime': instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(instance['LaunchTime'], datetime) else str(instance['LaunchTime']),
                    'PrivateIpAddress': instance.get('PrivateIpAddress'),
                    'PublicIpAddress': instance.get('PublicIpAddress'),
                    'VpcId': instance.get('VpcId'),
                    'SubnetId': instance.get('SubnetId'),
                    'AvailabilityZone': instance.get('Placement', {}).get('AvailabilityZone'),
                    'Tags': {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
                })
        resource_info['EC2Instances'] = all_instances
        print(f"  Encontradas {len(all_instances)} instâncias EC2.")
    except Exception as e:
        print(f"  Erro ao listar instâncias EC2: {e}")
        resource_info['EC2Instances'] = f"Erro ao listar: {e}"

    print("Coletando informações de VPCs...")
    try:
        vpcs = get_all_paginated_items(ec2_client, 'describe_vpcs', 'Vpcs')
        all_vpcs = []
        for vpc in vpcs:
            all_vpcs.append({
                'VpcId': vpc['VpcId'],
                'CidrBlock': vpc['CidrBlock'],
                'IsDefault': vpc['IsDefault'],
                'State': vpc['State'],
                'Tags': {tag['Key']: tag['Value'] for tag in vpc.get('Tags', [])}
            })
        resource_info['VPCs'] = all_vpcs
        print(f"  Encontradas {len(all_vpcs)} VPCs.")
    except Exception as e:
        print(f"  Erro ao listar VPCs: {e}")
        resource_info['VPCs'] = f"Erro ao listar: {e}"

    print("Coletando informações de Transit Gateways...")
    try:
        tgws = get_all_paginated_items(ec2_client, 'describe_transit_gateways', 'TransitGateways')
        all_tgws = []
        for tgw in tgws:
            all_tgws.append({
                'TransitGatewayId': tgw['TransitGatewayId'],
                'State': tgw['State'],
                'Description': tgw.get('Description', 'N/A'),
                'CreationTime': tgw['CreationTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(tgw['CreationTime'], datetime) else str(tgw['CreationTime']),
                'Tags': {tag['Key']: tag['Value'] for tag in tgw.get('Tags', [])}
            })
        resource_info['TransitGateways'] = all_tgws
        print(f"  Encontrados {len(all_tgws)} Transit Gateways.")
    except Exception as e:
        print(f"  Erro ao listar Transit Gateways: {e}")
        resource_info['TransitGateways'] = f"Erro ao listar: {e}"

    # --- S3 Buckets ---
    s3_client = boto3.client('s3')
    print("Coletando informações de buckets S3...")
    try:
        s3_response = s3_client.list_buckets() # S3 list_buckets não usa paginação no mesmo formato de NextToken
        all_s3_buckets = []
        for bucket in s3_response['Buckets']:
            all_s3_buckets.append({
                'Name': bucket['Name'],
                'CreationDate': bucket['CreationDate'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(bucket['CreationDate'], datetime) else str(bucket['CreationDate'])
            })
        resource_info['S3Buckets'] = all_s3_buckets
        print(f"  Encontrados {len(all_s3_buckets)} buckets S3.")
    except Exception as e:
        print(f"  Erro ao listar buckets S3: {e}")
        resource_info['S3Buckets'] = f"Erro ao listar: {e}"

    # --- Lambda Functions ---
    lambda_client = boto3.client('lambda')
    print("Coletando informações de funções Lambda...")
    try:
        functions = get_all_paginated_items(lambda_client, 'list_functions', 'Functions')
        all_lambda_functions = []
        for func in functions:
            all_lambda_functions.append({
                'FunctionName': func['FunctionName'],
                'Runtime': func['Runtime'],
                'MemorySize': func['MemorySize'],
                'LastModified': func['LastModified']
            })
        resource_info['LambdaFunctions'] = all_lambda_functions
        print(f"  Encontradas {len(all_lambda_functions)} funções Lambda.")
    except Exception as e:
        print(f"  Erro ao listar funções Lambda: {e}")
        resource_info['LambdaFunctions'] = f"Erro ao listar: {e}"

    # --- RDS/Aurora Databases ---
    rds_client = boto3.client('rds')
    
    print("Coletando informações de instâncias RDS (não-Aurora)...")
    try:
        # Instâncias RDS que NÃO fazem parte de clusters Aurora
        db_instances = get_all_paginated_items(rds_client, 'describe_db_instances', 'DBInstances')
        all_rds_instances = []
        for db in db_instances:
            # Filtra instâncias que são membros de um cluster DB (ex: Aurora)
            if not db.get('DBClusterIdentifier'):
                all_rds_instances.append({
                    'DBInstanceIdentifier': db['DBInstanceIdentifier'],
                    'Engine': db['Engine'],
                    'EngineVersion': db.get('EngineVersion'),
                    'DBInstanceClass': db['DBInstanceClass'],
                    'DBInstanceStatus': db['DBInstanceStatus'],
                    'AllocatedStorage': db['AllocatedStorage'],
                    'EndpointAddress': db.get('Endpoint', {}).get('Address'),
                    'MultiAZ': db['MultiAZ'],
                    'PubliclyAccessible': db['PubliclyAccessible'],
                    'BackupRetentionPeriod': db['BackupRetentionPeriod'],
                    'VpcSecurityGroupIds': [sg['VpcSecurityGroupId'] for sg in db.get('VpcSecurityGroups', [])],
                    'Tags': {tag['Key']: tag['Value'] for tag in db.get('TagList', [])}
                })
        resource_info['RDSInstances'] = all_rds_instances
        print(f"  Encontradas {len(all_rds_instances)} instâncias RDS (não-Aurora).")
    except Exception as e:
        print(f"  Erro ao listar instâncias RDS (não-Aurora): {e}")
        resource_info['RDSInstances'] = f"Erro ao listar: {e}"

    print("Coletando informações de clusters Aurora...")
    try:
        # Clusters Aurora
        db_clusters = get_all_paginated_items(rds_client, 'describe_db_clusters', 'DBClusters')
        all_aurora_clusters = []
        for cluster in db_clusters:
            # Lista as instâncias associadas ao cluster
            cluster_members = []
            for member in cluster.get('DBClusterMembers', []):
                cluster_members.append({
                    'DBInstanceIdentifier': member['DBInstanceIdentifier'],
                    'IsClusterWriter': member['IsClusterWriter'],
                    'PromotionTier': member.get('PromotionTier')
                })

            all_aurora_clusters.append({
                'DBClusterIdentifier': cluster['DBClusterIdentifier'],
                'Engine': cluster['Engine'],
                'EngineVersion': cluster['EngineVersion'],
                'Status': cluster['Status'],
                'Endpoint': cluster.get('Endpoint'),
                'ReaderEndpoint': cluster.get('ReaderEndpoint'),
                'MultiAZ': cluster['MultiAZ'],
                'BackupRetentionPeriod': cluster['BackupRetentionPeriod'],
                'AllocatedStorage': cluster.get('AllocatedStorage'), # Aurora storage can vary, this is an estimate
                'VpcSecurityGroupIds': [sg['VpcSecurityGroupId'] for sg in cluster.get('VpcSecurityGroups', [])],
                'ClusterMembers': cluster_members,
                'Tags': {tag['Key']: tag['Value'] for tag in cluster.get('TagList', [])}
            })
        resource_info['AuroraClusters'] = all_aurora_clusters
        print(f"  Encontrados {len(all_aurora_clusters)} clusters Aurora.")
    except Exception as e:
        print(f"  Erro ao listar clusters Aurora: {e}")
        resource_info['AuroraClusters'] = f"Erro ao listar: {e}"

    # --- Load Balancers (ALB/NLB/ELB) ---
    elbv2_client = boto3.client('elbv2') # Para ALB/NLB
    elb_client = boto3.client('elb')     # Para Classic Load Balancers
    print("Coletando informações de Load Balancers...")
    try:
        # ALB/NLB
        lbs_v2 = get_all_paginated_items(elbv2_client, 'describe_load_balancers', 'LoadBalancers')
        all_lbs_v2 = []
        for lb in lbs_v2:
            all_lbs_v2.append({
                'LoadBalancerArn': lb['LoadBalancerArn'],
                'LoadBalancerName': lb['LoadBalancerName'],
                'Type': lb['Type'],
                'State': lb['State']['Code'],
                'Scheme': lb['Scheme'],
                'VpcId': lb['VpcId'],
                'CreatedTime': lb['CreatedTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(lb['CreatedTime'], datetime) else str(lb['CreatedTime']),
            })
        resource_info['LoadBalancersV2'] = all_lbs_v2
        print(f"  Encontrados {len(all_lbs_v2)} Load Balancers (ALB/NLB).")

        # Classic Load Balancers
        lbs_classic = get_all_paginated_items(elb_client, 'describe_load_balancers', 'LoadBalancerDescriptions')
        all_lbs_classic = []
        for lb in lbs_classic:
            all_lbs_classic.append({
                'LoadBalancerName': lb['LoadBalancerName'],
                'Scheme': lb['Scheme'],
                'VPCId': lb.get('VPCId'),
                'CreatedTime': lb['CreatedTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(lb['CreatedTime'], datetime) else str(lb['CreatedTime']),
                'Instances': [inst['InstanceId'] for inst in lb.get('Instances', [])]
            })
        resource_info['ClassicLoadBalancers'] = all_lbs_classic
        print(f"  Encontrados {len(all_lbs_classic)} Classic Load Balancers.")

    except Exception as e:
        print(f"  Erro ao listar Load Balancers: {e}")
        resource_info['LoadBalancers'] = f"Erro ao listar: {e}"

    # --- Auto Scaling Groups ---
    autoscaling_client = boto3.client('autoscaling')
    print("Coletando informações de Auto Scaling Groups...")
    try:
        asgs = get_all_paginated_items(autoscaling_client, 'describe_auto_scaling_groups', 'AutoScalingGroups')
        all_asgs = []
        for asg in asgs:
            all_asgs.append({
                'AutoScalingGroupName': asg['AutoScalingGroupName'],
                'MinSize': asg['MinSize'],
                'MaxSize': asg['MaxSize'],
                'DesiredCapacity': asg['DesiredCapacity'],
                'HealthCheckType': asg['HealthCheckType'],
                'VPCZoneIdentifier': asg.get('VPCZoneIdentifier'),
                'Instances': [{'InstanceId': inst['InstanceId'], 'LifecycleState': inst['LifecycleState']} for inst in asg.get('Instances', [])]
            })
        resource_info['AutoScalingGroups'] = all_asgs
        print(f"  Encontrados {len(all_asgs)} Auto Scaling Groups.")
    except Exception as e:
        print(f"  Erro ao listar Auto Scaling Groups: {e}")
        resource_info['AutoScalingGroups'] = f"Erro ao listar: {e}"

    # --- IAM Roles, Users, Policies ---
    iam_client = boto3.client('iam')
    print("Coletando informações de IAM...")
    try:
        # Users
        users = get_all_paginated_items(iam_client, 'list_users', 'Users')
        all_users = []
        for user in users:
            all_users.append({
                'UserName': user['UserName'],
                'UserId': user['UserId'],
                'Arn': user['Arn'],
                'CreateDate': user['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(user['CreateDate'], datetime) else str(user['CreateDate'])
            })
        resource_info['IAMUsers'] = all_users
        print(f"    Encontrados {len(all_users)} usuários IAM.")

        # Roles
        roles = get_all_paginated_items(iam_client, 'list_roles', 'Roles')
        all_roles = []
        for role in roles:
            all_roles.append({
                'RoleName': role['RoleName'],
                'RoleId': role['RoleId'],
                'Arn': role['Arn'],
                'CreateDate': role['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(role['CreateDate'], datetime) else str(role['CreateDate']),
                'AssumeRolePolicyDocument': json.dumps(role.get('AssumeRolePolicyDocument')) # Convert policy to string
            })
        resource_info['IAMRoles'] = all_roles
        print(f"    Encontradas {len(all_roles)} roles IAM.")

        # Policies (Customer Managed Policies only)
        policies = get_all_paginated_items(iam_client, 'list_policies', 'Policies', Scope='Local')
        all_policies = []
        for policy in policies:
            all_policies.append({
                'PolicyName': policy['PolicyName'],
                'PolicyId': policy['PolicyId'],
                'Arn': policy['Arn'],
                'CreateDate': policy['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(policy['CreateDate'], datetime) else str(policy['CreateDate']),
                'AttachmentCount': policy['AttachmentCount']
            })
        resource_info['IAMPolicies'] = all_policies
        print(f"    Encontradas {len(all_policies)} políticas IAM (customizadas).")

    except Exception as e:
        print(f"  Erro ao listar recursos IAM: {e}")
        resource_info['IAMResources'] = f"Erro ao listar: {e}"

    # --- CloudFormation Stacks ---
    cf_client = boto3.client('cloudformation')
    print("Coletando informações de CloudFormation Stacks...")
    try:
        stacks = get_all_paginated_items(cf_client, 'describe_stacks', 'Stacks')
        all_stacks = []
        for stack in stacks:
            all_stacks.append({
                'StackName': stack['StackName'],
                'StackId': stack['StackId'],
                'StackStatus': stack['StackStatus'],
                'CreationTime': stack['CreationTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(stack['CreationTime'], datetime) else str(stack['CreationTime']),
                'LastUpdatedTime': stack['LastUpdatedTime'].strftime('%Y-%m-%d %H:%M:%S') if stack.get('LastUpdatedTime') and isinstance(stack['LastUpdatedTime'], datetime) else str(stack.get('LastUpdatedTime', 'N/A')),
                'Outputs': stack.get('Outputs', [])
            })
        resource_info['CloudFormationStacks'] = all_stacks
        print(f"  Encontradas {len(all_stacks)} CloudFormation Stacks.")
    except Exception as e:
        print(f"  Erro ao listar CloudFormation Stacks: {e}")
        resource_info['CloudFormationStacks'] = f"Erro ao listar: {e}"

    # --- DynamoDB Tables ---
    dynamodb_client = boto3.client('dynamodb')
    print("Coletando informações de tabelas DynamoDB...")
    try:
        table_names = get_all_paginated_items(dynamodb_client, 'list_tables', 'TableNames')
        all_dynamodb_tables = []
        for table_name in table_names:
            # Para obter detalhes, precisamos chamar describe_table para cada tabela
            table_desc = dynamodb_client.describe_table(TableName=table_name)['Table']
            all_dynamodb_tables.append({
                'TableName': table_desc['TableName'],
                'TableStatus': table_desc['TableStatus'],
                'CreationDateTime': table_desc['CreationDateTime'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(table_desc['CreationDateTime'], datetime) else str(table_desc['CreationDateTime']),
                'TableSizeBytes': table_desc.get('TableSizeBytes'),
                'ItemCount': table_desc.get('ItemCount'),
                'ProvisionedThroughput': {
                    'ReadCapacityUnits': table_desc['ProvisionedThroughput']['ReadCapacityUnits'],
                    'WriteCapacityUnits': table_desc['ProvisionedThroughput']['WriteCapacityUnits']
                },
                'BillingModeSummary': table_desc.get('BillingModeSummary', {}).get('BillingMode')
            })
        resource_info['DynamoDBTables'] = all_dynamodb_tables
        print(f"  Encontradas {len(all_dynamodb_tables)} tabelas DynamoDB.")
    except Exception as e:
        print(f"  Erro ao listar tabelas DynamoDB: {e}")
        resource_info['DynamoDBTables'] = f"Erro ao listar: {e}"

    # --- SNS Topics ---
    sns_client = boto3.client('sns')
    print("Coletando informações de tópicos SNS...")
    try:
        topics = get_all_paginated_items(sns_client, 'list_topics', 'Topics')
        all_sns_topics = []
        for topic in topics:
            all_sns_topics.append({
                'TopicArn': topic['TopicArn']
                # Mais detalhes podem ser obtidos com get_topic_attributes se necessário
            })
        resource_info['SNSTopics'] = all_sns_topics
        print(f"  Encontrados {len(all_sns_topics)} tópicos SNS.")
    except Exception as e:
        print(f"  Erro ao listar tópicos SNS: {e}")
        resource_info['SNSTopics'] = f"Erro ao listar: {e}"

    # --- SQS Queues ---
    sqs_client = boto3.client('sqs')
    print("Coletando informações de filas SQS...")
    try:
        queue_urls = get_all_paginated_items(sqs_client, 'list_queues', 'QueueUrls')
        all_sqs_queues = []
        for queue_url in queue_urls:
            # Para obter detalhes, pode-se usar get_queue_attributes
            attributes = sqs_client.get_queue_attributes(
                QueueUrl=queue_url,
                AttributeNames=['ApproximateNumberOfMessages', 'CreatedTimestamp', 'LastModifiedTimestamp']
            )['Attributes']
            all_sqs_queues.append({
                'QueueUrl': queue_url,
                'ApproximateNumberOfMessages': attributes.get('ApproximateNumberOfMessages'),
                'CreatedTimestamp': datetime.fromtimestamp(int(attributes['CreatedTimestamp'])).strftime('%Y-%m-%d %H:%M:%S'),
                'LastModifiedTimestamp': datetime.fromtimestamp(int(attributes['LastModifiedTimestamp'])).strftime('%Y-%m-%d %H:%M:%S')
            })
        resource_info['SQSQueues'] = all_sqs_queues
        print(f"  Encontradas {len(all_sqs_queues)} filas SQS.")
    except Exception as e:
        print(f"  Erro ao listar filas SQS: {e}")
        resource_info['SQSQueues'] = f"Erro ao listar: {e}"

    # --- EKS Clusters ---
    eks_client = boto3.client('eks')
    print("Coletando informações de clusters EKS...")
    try:
        cluster_names = get_all_paginated_items(eks_client, 'list_clusters', 'clusters')
        all_eks_clusters = []
        for cluster_name in cluster_names:
            cluster_desc = eks_client.describe_cluster(name=cluster_name)['cluster']
            all_eks_clusters.append({
                'ClusterName': cluster_desc['name'],
                'Status': cluster_desc['status'],
                'Version': cluster_desc['version'],
                'Arn': cluster_desc['arn'],
                'CreatedAt': cluster_desc['createdAt'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(cluster_desc['createdAt'], datetime) else str(cluster_desc['createdAt']),
                'VpcConfig': cluster_desc.get('resourcesVpcConfig')
            })
        resource_info['EKSClusters'] = all_eks_clusters
        print(f"  Encontrados {len(all_eks_clusters)} clusters EKS.")
    except Exception as e:
        print(f"  Erro ao listar clusters EKS: {e}")
        resource_info['EKSClusters'] = f"Erro ao listar: {e}"

    # --- ECS Clusters and Services ---
    ecs_client = boto3.client('ecs')
    print("Coletando informações de clusters e serviços ECS...")
    try:
        cluster_arns = get_all_paginated_items(ecs_client, 'list_clusters', 'clusterArns')
        all_ecs_clusters = []
        for cluster_arn in cluster_arns:
            cluster_desc = ecs_client.describe_clusters(clusters=[cluster_arn])['clusters'][0]
            all_ecs_clusters.append({
                'ClusterName': cluster_desc['clusterName'],
                'ClusterArn': cluster_desc['clusterArn'],
                'Status': cluster_desc['status'],
                'RunningTasksCount': cluster_desc['runningTasksCount'],
                'RegisteredContainerInstancesCount': cluster_desc['registeredContainerInstancesCount'],
            })

            # Listar serviços dentro de cada cluster ECS
            service_arns = get_all_paginated_items(ecs_client, 'list_services', 'serviceArns', cluster=cluster_arn)
            all_ecs_services = []
            if service_arns:
                services_desc = ecs_client.describe_services(cluster=cluster_arn, services=service_arns)['services']
                for service in services_desc:
                    all_ecs_services.append({
                        'ServiceName': service['serviceName'],
                        'ServiceArn': service['serviceArn'],
                        'Status': service['status'],
                        'DesiredCount': service['desiredCount'],
                        'RunningCount': service['runningCount'],
                        'LaunchType': service.get('launchType'),
                        'CreatedAt': service['createdAt'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(service['createdAt'], datetime) else str(service['createdAt']),
                    })
            all_ecs_clusters[-1]['Services'] = all_ecs_services # Adiciona serviços ao cluster pai

        resource_info['ECSClusters'] = all_ecs_clusters
        print(f"  Encontrados {len(all_ecs_clusters)} clusters ECS com seus serviços.")
    except Exception as e:
        print(f"  Erro ao listar clusters e serviços ECS: {e}")
        resource_info['ECSClusters'] = f"Erro ao listar: {e}"

    return resource_info

if __name__ == "__main__":
    print("Iniciando a coleta de informações de recursos AWS...\n")
    aws_resources = get_aws_resources_info()
    
    # Imprime as informações em formato JSON para facilitar a leitura
    print("\n--- Resumo Detalhado dos Recursos AWS (JSON) ---")
    print(json.dumps(aws_resources, indent=2, default=str)) 
    
    # Exporta para Excel, usando "resources.xlsx" como nome padrão
    export_to_excel(aws_resources, filename="resources.xlsx")
    
    print("\nColeta e exportação de informações concluídas.")